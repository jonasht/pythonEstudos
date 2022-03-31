import openpyxl
book = openpyxl.Workbook()

# print(book.sheetnames)

# criar uma pagina
book.create_sheet('frutas')

# selecionar uma pagina
frutas_page = book['frutas']
frutas_page.append(['fruta', 'quantidade', 'preço'])
frutas_page.append(['maça', '5', '3,22'])
frutas_page.append(['banana', '5', '3,22'])
frutas_page.append(['banana', '5', '3,22'])
frutas_page.append(['banana', '5', '3,22'])
frutas_page.append(['banana', '5', '3,22'])
frutas_page.append(['banana', '5', '3,22'])
frutas_page.append(['banana', '5', '3,22'])
frutas_page.append(['banana', '5', '3,22'])

book.save('planilhas.xlsx')
import openpyxl


book = openpyxl.load_workbook('planilhas.xlsx')

frutas_page = book['frutas']

# for row in frutas_page.iter_rows(min_col=2, max_col=5)
for rows in frutas_page.iter_rows():
    for w in rows:
        print(f'{w.value:<10} ', end='')
    print()
